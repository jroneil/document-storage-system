import csv
import json
import xmltodict
import pandas as pd
import hashlib
import uuid
from typing import List, Dict, Any
from datetime import datetime
import logging
from openpyxl import load_workbook

# Configure logging
logger = logging.getLogger(__name__)

class MetadataParser:
    def __init__(self):
        self.schema_version = "v1"
        self.batch_size = 500  # Default batch size
    
    def parse_file(self, file_path: str, job_id: str = None) -> Dict[str, Any]:
        """Parse file and return normalized records with job information"""
        if not job_id:
            job_id = str(uuid.uuid4())
        
        file_extension = file_path.lower().split('.')[-1]
        
        try:
            if file_extension == 'csv':
                raw_records = self._parse_csv(file_path)
            elif file_extension == 'json':
                raw_records = self._parse_json(file_path)
            elif file_extension == 'xml':
                raw_records = self._parse_xml(file_path)
            elif file_extension in ['xlsx', 'xls']:
                raw_records = self._parse_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            # Normalize and validate records
            normalized_records = []
            validation_errors = []
            
            for i, raw_record in enumerate(raw_records):
                try:
                    normalized_record = self._normalize_record(raw_record, job_id, i + 2)  # +2 for header row
                    normalized_records.append(normalized_record)
                except Exception as e:
                    validation_errors.append({
                        "row_num": i + 2,
                        "error": str(e),
                        "raw_data": raw_record
                    })
                    logger.warning(f"Validation error in row {i + 2}: {e}")
            
            # Create batch message
            batch_message = {
                "job_id": job_id,
                "batch_id": 1,
                "schema_version": self.schema_version,
                "source_file": file_path.split('/')[-1],
                "records": normalized_records,
                "validation_errors": validation_errors
            }
            
            return batch_message
            
        except Exception as e:
            logger.error(f"Failed to parse file {file_path}: {e}")
            raise
    
    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse CSV file using streaming approach"""
        records = []
        with open(file_path, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                records.append(row)
        return records
    
    def _parse_json(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse JSON file"""
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Handle both array of objects and single object
        if isinstance(data, list):
            return data
        elif isinstance(data, dict):
            return [data]
        else:
            raise ValueError("JSON file must contain an array of objects or a single object")
    
    def _parse_xml(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse XML file"""
        with open(file_path, 'r', encoding='utf-8') as file:
            xml_content = file.read()
        
        # Parse XML to dictionary
        data = xmltodict.parse(xml_content)
        
        # Extract records - assuming XML structure with records array
        records = []
        if 'records' in data and 'record' in data['records']:
            records_data = data['records']['record']
            if isinstance(records_data, list):
                records = records_data
            else:
                records = [records_data]
        elif 'record' in data:
            records_data = data['record']
            if isinstance(records_data, list):
                records = records_data
            else:
                records = [records_data]
        else:
            # Try to find any array of objects
            for key, value in data.items():
                if isinstance(value, list):
                    records = value
                    break
            else:
                records = [data]  # Single record
        
        return records
    
    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """Parse Excel file"""
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = [str(cell.value) for cell in sheet[1]]  # Read the header row
        
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i, cell_value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = cell_value
            records.append(row_data)
        
        return records
    
    def _normalize_record(self, raw_record: Dict[str, Any], job_id: str, row_num: int) -> Dict[str, Any]:
        """Normalize and validate a single record according to canonical schema"""
        
        # Extract and validate required fields
        file_name = self._get_required_field(raw_record, 'file_name')
        file_size = self._get_required_field(raw_record, 'file_size')
        file_type = self._get_required_field(raw_record, 'file_type')
        user_id = self._get_required_field(raw_record, 'user_id')
        storage_path = self._get_required_field(raw_record, 'storage_path')
        version = self._get_required_field(raw_record, 'version')
        checksum = self._get_required_field(raw_record, 'checksum')
        acl = self._get_required_field(raw_record, 'acl')
        
        # Validate types and constraints
        file_size = self._validate_file_size(file_size)
        version = self._validate_version(version)
        acl_data = self._validate_acl(acl)
        
        # Compute record_id
        record_id = self._compute_record_id(file_name, checksum, storage_path)
        
        # Build normalized record
        normalized_record = {
            "record_id": record_id,
            "source": {
                "job_id": job_id,
                "batch_id": 1,
                "row_num": row_num,
                "filename": raw_record.get('filename', 'unknown'),
                "schema_version": self.schema_version
            },
            "asset": {
                "file_name": file_name,
                "file_size": file_size,
                "file_type": file_type,
                "version": version,
                "checksum": checksum,
                "storage_path": storage_path,
                "thumbnail_path": self._get_optional_field(raw_record, 'thumbnail_path'),
                "expiration_date": self._validate_date(self._get_optional_field(raw_record, 'expiration_date'))
            },
            "ownership": {
                "uploader_user_id": user_id,
                "acl": acl_data
            },
            "metadata": {
                "tags": self._parse_list_field(raw_record.get('tags', '')),
                "description": self._get_optional_field(raw_record, 'description'),
                "category": self._get_optional_field(raw_record, 'category'),
                "division": self._get_optional_field(raw_record, 'division'),
                "business_unit": self._get_optional_field(raw_record, 'business_unit'),
                "brand_id": self._get_optional_field(raw_record, 'brand_id'),
                "document_type": self._get_optional_field(raw_record, 'document_type'),
                "region": self._get_optional_field(raw_record, 'region'),
                "country": self._get_optional_field(raw_record, 'country'),
                "languages": self._parse_languages(raw_record.get('languages', '')),
                "alternate_part_numbers": self._parse_list_field(raw_record.get('alternate_part_numbers', ''))
            }
        }
        
        return normalized_record
    
    def _get_required_field(self, record: Dict[str, Any], field_name: str) -> Any:
        """Get required field or raise validation error"""
        value = record.get(field_name)
        if value is None or value == '':
            raise ValueError(f"Required field '{field_name}' is missing or empty")
        return value
    
    def _get_optional_field(self, record: Dict[str, Any], field_name: str) -> Any:
        """Get optional field, return None if not present or empty"""
        value = record.get(field_name)
        if value is None or value == '':
            return None
        return value
    
    def _validate_file_size(self, file_size: Any) -> int:
        """Validate and convert file_size to integer"""
        try:
            size = int(file_size)
            if size < 0:
                raise ValueError("file_size must be >= 0")
            return size
        except (ValueError, TypeError):
            raise ValueError(f"Invalid file_size: {file_size}")
    
    def _validate_version(self, version: Any) -> int:
        """Validate and convert version to integer"""
        try:
            ver = int(version)
            if ver < 1:
                raise ValueError("version must be >= 1")
            return ver
        except (ValueError, TypeError):
            raise ValueError(f"Invalid version: {version}")
    
    def _validate_acl(self, acl: Any) -> Dict[str, List[str]]:
        """Validate and parse ACL field"""
        if isinstance(acl, dict):
            # Already parsed as dict
            acl_data = acl
        elif isinstance(acl, str):
            # Parse JSON string
            try:
                acl_data = json.loads(acl)
            except json.JSONDecodeError:
                raise ValueError(f"Invalid ACL JSON: {acl}")
        else:
            raise ValueError(f"Invalid ACL format: {acl}")
        
        # Ensure required structure
        if not isinstance(acl_data, dict):
            raise ValueError("ACL must be a dictionary")
        
        result = {"read": [], "write": []}
        
        if 'read' in acl_data:
            if isinstance(acl_data['read'], list):
                result['read'] = [str(item) for item in acl_data['read']]
            else:
                raise ValueError("ACL 'read' must be an array")
        
        if 'write' in acl_data:
            if isinstance(acl_data['write'], list):
                result['write'] = [str(item) for item in acl_data['write']]
            else:
                raise ValueError("ACL 'write' must be an array")
        
        return result
    
    def _validate_date(self, date_str: str) -> str:
        """Validate date format (YYYY-MM-DD)"""
        if not date_str:
            return None
        
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            raise ValueError(f"Invalid date format: {date_str}. Expected YYYY-MM-DD")
    
    def _parse_list_field(self, field_value: Any) -> List[str]:
        """Parse comma-separated list field"""
        if not field_value:
            return []
        
        if isinstance(field_value, list):
            return [str(item).strip() for item in field_value if str(item).strip()]
        
        if isinstance(field_value, str):
            items = [item.strip() for item in field_value.split(',') if item.strip()]
            return items
        
        return [str(field_value).strip()]
    
    def _parse_languages(self, languages: Any) -> List[str]:
        """Parse languages field and convert to lowercase ISO codes"""
        languages_list = self._parse_list_field(languages)
        return [lang.strip().lower() for lang in languages_list if lang.strip()]
    
    def _compute_record_id(self, file_name: str, checksum: str, storage_path: str) -> str:
        """Compute deterministic record_id using SHA256"""
        key = f"{file_name.lower()}|{checksum}|{storage_path}"
        return hashlib.sha256(key.encode('utf-8')).hexdigest()


# Backward compatibility function
def parse_metadata(file_path: str) -> List[Dict[str, Any]]:
    """Legacy function for backward compatibility"""
    parser = MetadataParser()
    result = parser.parse_file(file_path)
    return result.get('records', [])


# Export the parser class
__all__ = ['MetadataParser', 'parse_metadata']
